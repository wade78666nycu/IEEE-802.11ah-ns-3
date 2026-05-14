#!/usr/bin/env python3
"""Convert comparison_report.txt to a formatted xlsx file."""

import sys
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

REPORT = Path(__file__).parent / "output_file" / "comparison_report.txt"
OUTPUT = Path(__file__).parent / "output_file" / "comparison_report.xlsx"

HEADERS = ["Seed", "Traffic", "Case", "PDR(%)", "Thru(Kbps)", "AvgDly(ms)", "RERR", "Energy(J)", "E/Byte(mJ)"]

# colours
HDR_FILL  = PatternFill("solid", fgColor="2E4057")   # dark blue
LIGHT_FILL = PatternFill("solid", fgColor="EBF5FB")  # light blue
HEAVY_FILL = PatternFill("solid", fgColor="FEF9E7")  # light yellow
ALT_FILLS  = {"light": LIGHT_FILL, "heavy": HEAVY_FILL}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def parse_report(path: Path) -> list[dict]:
    rows = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("─"):
            continue
        # skip the header line
        if line.startswith("Seed"):
            continue
        parts = line.split()
        if len(parts) < 8:
            continue
        try:
            rows.append({
                "Seed":       parts[0],
                "Traffic":    parts[1],
                "Case":       parts[2],
                "PDR(%)":     float(parts[3]),
                "Thru(Kbps)": float(parts[4]),
                "AvgDly(ms)": float(parts[5]),
                "RERR":       int(parts[6]),
                "Energy(J)":  float(parts[7]),
                "E/Byte(mJ)": float(parts[8]) if len(parts) > 8 else "N/A",
            })
        except (ValueError, IndexError):
            continue
    return rows


def write_xlsx(rows: list[dict], output: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # header row
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    # data rows
    for r, row in enumerate(rows, 2):
        fill = ALT_FILLS.get(row["Traffic"], LIGHT_FILL)
        for col, key in enumerate(HEADERS, 1):
            val = row[key]
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left")
            if isinstance(val, float):
                if key == "Energy(J)":    cell.number_format = "0.0000"
                elif key == "E/Byte(mJ)": cell.number_format = "0.000000"
                else:                      cell.number_format = "0.00"

    # column widths
    widths = [6, 8, 26, 9, 12, 12, 7, 11, 13]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # freeze header
    ws.freeze_panes = "A2"

    wb.save(output)
    print(f"Saved: {output}")


if __name__ == "__main__":
    report = Path(sys.argv[1]) if len(sys.argv) > 1 else REPORT
    out    = report.with_suffix(".xlsx")
    rows = parse_report(report)
    if not rows:
        print(f"No data found in {report}")
        sys.exit(1)
    write_xlsx(rows, out)
