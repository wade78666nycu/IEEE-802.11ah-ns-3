#!/usr/bin/env python3
"""Convert grid_comparison_report.txt to a formatted xlsx file.

Mirrors report_to_xlsx.py (used by run_comparison.sh for rand) but for the grid
spatial-reuse sweep columns: Spacing / Rate / Mode / PDR / Thru / AvgDly / RERR /
Energy. Rows are coloured by Mode (full vs gradpc) so the two are easy to compare.
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Run: pip install openpyxl")
    sys.exit(1)

REPORT = Path(__file__).parent / "output_file" / "grid_comparison_report.txt"

HEADERS = ["Spacing", "Rate", "Mode", "PDR(%)", "Thru(Kbps)", "AvgDly(ms)", "RERR", "Energy(J)"]

HDR_FILL   = PatternFill("solid", fgColor="2E4057")   # dark blue
FULL_FILL  = PatternFill("solid", fgColor="EBF5FB")   # light blue  -> full power
GRAD_FILL  = PatternFill("solid", fgColor="E9F7EF")   # light green -> gradpc
MODE_FILLS = {"full": FULL_FILL, "gradpc": GRAD_FILL}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def parse_report(path: Path) -> list[dict]:
    rows = []
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("─") or line.startswith("Spacing"):
            continue
        parts = line.split()
        if len(parts) < 8:
            continue
        try:
            rows.append({
                "Spacing":    int(parts[0]),
                "Rate":       parts[1],
                "Mode":       parts[2],
                "PDR(%)":     float(parts[3]),
                "Thru(Kbps)": float(parts[4]),
                "AvgDly(ms)": float(parts[5]),
                "RERR":       int(parts[6]),
                "Energy(J)":  float(parts[7]),
            })
        except (ValueError, IndexError):
            continue
    return rows


def write_xlsx(rows: list[dict], output: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grid Comparison"

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    for r, row in enumerate(rows, 2):
        fill = MODE_FILLS.get(row["Mode"], FULL_FILL)
        for col, key in enumerate(HEADERS, 1):
            val = row[key]
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if isinstance(val, float):
                cell.number_format = "0.0000" if key == "Energy(J)" else "0.00"
            if key == "Mode" and row["Mode"] == "gradpc":
                cell.font = Font(bold=True)

    widths = [9, 9, 9, 9, 12, 12, 7, 11]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    wb.save(output)
    print(f"Saved: {output}")


if __name__ == "__main__":
    report = Path(sys.argv[1]) if len(sys.argv) > 1 else REPORT
    out = report.with_suffix(".xlsx")
    rows = parse_report(report)
    if not rows:
        print(f"No data found in {report}")
        sys.exit(1)
    write_xlsx(rows, out)
